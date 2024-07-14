import os
import csv
from openpyxl import load_workbook

def find_keywords_by_attrs(directory, attrs):
    keyword_counts = {func_name: 0 for func_name in attrs}

    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory, filename)
            workbook = load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                keywords = row[2]
                if keywords:
                    for keyword in keywords.split(', '):
                        for func_name, func_keywords in attrs.items():
                            if any(partial_word.lower() in keyword.lower() for partial_word in func_keywords):
                                keyword_counts[func_name] += 1

    with open('found-keywords/attrs_keyword_counts.csv', 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Obj', 'Count'])
        for func_name, count in keyword_counts.items():
            writer.writerow([func_name, count])

if __name__ == '__main__':
    attrs = {
        'Low Cost': ['low cost', 'affordable', 'inexpensive'],
        'SWaP': ['size weight power', 'swap', 'compact'],
        'High Reliability': ['high reliability', 'reliable', 'robust'],
        'Data Transfer Rate': ['data transfer rate', 'high speed', 'fast transfer'],
        'Energy Efficiency': ['energy efficient', 'low power', 'power saving'],
        'Scalability': ['scalable', 'scalability', 'expandable'],
        'Ease of Integration': ['easy integration', 'integratable', 'integration'],
        'Security': ['secure', 'security', 'safe'],
        'Interference Resistance': ['interference resistant', 'anti-interference', 'jamming resistant'],
        'Ease of Use': ['easy to use', 'user friendly', 'simple']
    }

    os.makedirs('found-keywords', exist_ok=True)
    find_keywords_by_attrs('practice-parsing/output', attrs)
    print('Поиск ключевых слов завершен успешно! Результаты сохранены в found-keywords/attrs_keyword_counts.csv')
