import os
import csv
from openpyxl import load_workbook

def find_keywords_by_attrs(directory, attrs):
    keyword_counts = {func_name: 0 for func_name in attrs}

    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory, filename)
            workbook = load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                keywords = row[2]
                if keywords:
                    for keyword in keywords.split(', '):
                        for func_name, func_keywords in attrs.items():
                            if any(partial_word.lower() in keyword.lower() for partial_word in func_keywords):
                                keyword_counts[func_name] += 1

    with open('found-keywords/output/domains_keyword_counts.csv', 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Obj', 'Count'])
        for func_name, count in keyword_counts.items():
            writer.writerow([func_name, count])

if __name__ == '__main__':
    application_domains = {
        'AI': ['ai', 'AI'],
        'Air Defense': ['air defense', 'anti-aircraft', 'aerial defense', 'air'],
        '5G/6G': ['5G', '6G', 'mobile network', 'next-gen communication'],
        'Maritime Navigation': ['maritime navigation', 'ship navigation', 'marine navigation', 'nautical navigation', 'maritime'],
        'Space Communication': ['space communication', 'satellite communication', 'space comm', 'space telecom'],
        'Autonomous Vehicles': ['autonomous vehicle', 'self-driving', 'autonomous car', 'driverless', 'unmanned vehicle'],
        'Internet of Things (IoT)': ['internet of things', 'IoT', 'connected devices', 'smart devices'],
        'Geological Exploration': ['geological exploration', 'mineral exploration', 'earth monitoring', 'geology survey', 'geological', 'exploration'],
        'Radar X-Band': ['x-band radar', 'x-band'],
        'Satellite Communication': ['satellite communication', 'satcomm', 'satellite comm', 'satellite telecom'],
        'Free-Space Optical Communications (FSO)': ['free-space optical', 'FSO', 'optical communication', 'laser communication'],
        'Phased Array Antennas': ['phased array antenna', 'phased array', 'PAA', 'antenna array'],
        'Monitoring': ['monitor', 'monitoring'],
        'Vector Network Analyzers, VNAs': ['vector network', 'VNAs', 'vector network analyzers', 'vnas'],
        'ISR, Intelligence, Surveillance, and Reconnaissance': ['isr, intelligence, surveillance, reconnaissance'],
        'CBRN, Chemical, Biological, Radiological, and Nuclear': ['cbrn', 'nuclear', 'chemical'],
        'PMCW, Phase Modulated Continuous Wave radars': ['pmcw', 'phase modulated'],
        'Field RF Analyzer': ['field rf', 'field analyze', 'field radio']

    }
    os.makedirs('found-keywords', exist_ok=True)
    find_keywords_by_attrs('practice-parsing/output', application_domains)
    print('Поиск ключевых слов завершен успешно! Результаты сохранены в found-keywords/output/domains_keyword_counts.csv')
