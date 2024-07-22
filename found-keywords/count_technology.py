import os
import csv
from openpyxl import load_workbook

def find_keywords_by_attrs(directory, attrs):
    keyword_counts = {func_name: 0 for func_name in attrs}

    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory, filename)
            workbook = load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                keywords = row[2]
                if keywords:
                    for keyword in keywords.split(', '):
                        for func_name, func_keywords in attrs.items():
                            if any(partial_word.lower() in keyword.lower() for partial_word in func_keywords):
                                keyword_counts[func_name] += 1

    with open('found-keywords/output/technology_keyword_counts.csv', 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Obj', 'Count'])
        for func_name, count in keyword_counts.items():
            writer.writerow([func_name, count])

if __name__ == '__main__':
    attrs = {
        'Machine Learning': ['machine learning', 'ML', 'artificial intelligence', 'AI', 'deep learning', 'neural network', 'data mining'],
        'FPGA': ['FPGA', 'field-programmable gate array', 'reconfigurable hardware', 'programmable logic', 'programmable'],
        'GaN (Gallium Nitride)': ['GaN', 'gallium nitride', 'wide bandgap semiconductor', 'high electron mobility transistor', 'HEMT'],
        'SiC (Silicon Carbide)': ['SiC', 'silicon carbide', 'wide bandgap semiconductor', 'power electronics'],
        'MMIC (Monolithic Microwave Integrated Circuit)': ['MMIC', 'monolithic microwave integrated circuit', 'microwave IC', 'RF IC', 'radio frequency integrated circuit'],
        'Digital Signal Processing (DSP)': ['DSP', 'digital signal processing', 'signal processing', 'digital filter', 'FFT', 'Fast Fourier Transform'],
        'Quantum Radars': ['quantum radar', 'quantum sensing', 'quantum entanglement', 'quantum technology'],
        'High-Speed ADC': ['high-speed ADC', 'analog-to-digital converter', 'ADC', 'data converter', 'high resolution ADC'],
        'High-Speed Data Buses for Sensor Fusion': ['high-speed data bus', 'sensor fusion', 'data integration', 'data bus', 'real-time data processing'],
        'Mesh Networks': ['mesh', 'mesh networks'],
        'mmWave Chip': ['mmwave', 'mmwave chip'],
        'Rad-Hard Technologies': ['rad-hard', 'rad-hard technologies'],
        'Intermittent Sampling': ['intermittent sampling', 'intermittent'],
        'SiGe, Silicon-Germanium': ['sige', 'silicon-germaninum', 'germaninum'],
        'OFDM, Orthogonal Frequency-Division Multiplexing': ['ofdm', 'frequency-division', 'orthogonal']

    }

    os.makedirs('found-keywords', exist_ok=True)
    find_keywords_by_attrs('practice-parsing/output', attrs)
    print('Поиск ключевых слов завершен успешно! Результаты сохранены в found-keywords/output/technology_keyword_counts.csv')
