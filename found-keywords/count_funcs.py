import os
import csv
from openpyxl import load_workbook

def find_keywords_by_function(directory, functions):
    keyword_counts = {func_name: 0 for func_name in functions}

    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory, filename)
            workbook = load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                keywords = row[2]
                if keywords:
                    for keyword in keywords.split(', '):
                        for func_name, func_keywords in functions.items():
                            if any(partial_word.lower() in keyword.lower() for partial_word in func_keywords):
                                keyword_counts[func_name] += 1

    with open('found-keywords/output/function_keyword_counts.csv', 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Obj', 'Count'])
        for func_name, count in keyword_counts.items():
            writer.writerow([func_name, count])

if __name__ == '__main__':
    functions = {
        'Amplification': ['amplify', 'amplificate', 'amplification'],
        'Coordination': ['coordination', 'coordinate'],
        'Data Transmission': ['transmit', 'transmission', 'data transmission'],
        'Satellite navigation': ['satellite', 'navigation', 'gnss', 'GNSS'],
        'Real-time Communication': ['real-time', 'real time', 'real-time communication'],
        'Full-duplex Communication': ['duplex', 'full-duplex'],
        'RFID': ['rfid', 'RFID', 'identification'],
        'Multi-target Tracking': ['multi-target', 'multi tracking'],
        'Noise Measurement': ['noise', 'noise measurement'],
        'Power Measurements': ['power measurements', 'measurements'],
        'Automotive Radar Testing Systems': ['adas', 'automotive radar'],
        'Spectrum Analyzers for Interference Detection': ['spectrum analyzers', 'interference detection'],
        'RF Filtering and Impedance Matching': ['impedance matching', 'impedance'],
        'Detection and classification of mines': ['mines', 'detection mines', 'detect mines', 'mine'],
    }

    os.makedirs('found-keywords', exist_ok=True)
    find_keywords_by_function('practice-parsing/output', functions)
    print('Поиск ключевых слов завершен успешно! Результаты сохранены в found-keywords/function_keyword_counts.csv')
