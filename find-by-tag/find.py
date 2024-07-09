import os
from openpyxl import load_workbook

def find_entities_by_tag(directory, tag='ORG'):
    found_entities = set()

    for filename in os.listdir('practice-parsing/output'):
            if filename.endswith('.xlsx'):
                file_path = os.path.join('practice-parsing/output', filename)
                workbook = load_workbook(file_path)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    entities = row[2]
                    if entities:
                        for entity in entities.split(', '):
                            entity_name, entity_tag = entity[:entity.find('(')], entity[entity.find('(')+1:entity.find(')')]
                            if entity_tag == 'ORG':
                                found_entities.add(entity_name)

    with open(f'find-by-tag/found-entities/{tag.lower()}_entities.txt', 'w', encoding='utf-8') as f:
        for entity in found_entities:
                f.write(entity + '\n') 

if __name__ == '__main__':
    find_entities_by_tag('practice-parsing/output', 'PRODUCT')
    print('Эскпорт значений по тегу завершен успешно! Он сохранен в директорию found-entities')
